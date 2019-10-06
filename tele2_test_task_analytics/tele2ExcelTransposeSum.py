import openpyxl, numpy
from openpyxl.utils import get_column_letter


input_excel_filename = 'input_table.xlsx'
output_excel_filename = 'output_table.xlsx'

# open and transpose existing table
excel = openpyxl.load_workbook(input_excel_filename)
sheet = excel.active
table = [cell for cell in sheet.values]
transposedTable = numpy.transpose(table).tolist()

# Assign the transposed list to the excel object and save
output_excel = openpyxl.Workbook()
output_sheet = output_excel.active
for row in transposedTable:
    output_sheet.append(row)

# Calculate sum of columns by Excel embedded functions
row_num = output_sheet.max_row + 1
for col_num in range(1, output_sheet.max_column + 1):
    col_letter = get_column_letter(col_num)
    cell_value = '=SUM(' + col_letter + '1:' + col_letter + str(row_num - 1) + ')'
    output_sheet.cell(row=row_num, column=col_num, value=cell_value)
    # cell_value)

# save transposed and calculated table to the new file
output_excel.save(output_excel_filename)
